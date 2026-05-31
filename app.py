import os
import datetime
import threading
import click

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_migrate import Migrate
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from software_services.User_services import UserService
from software_services.clinic_services import ClinicService 
from software_services.complaint_services import ComplaintService
from software_services.doctor_services import DoctorService  
from software_services.specialty_services import SpecialtyService
from software_services.service_services import ServiceService
from software_services.platform_services import PlatformService
from software_services.page_services import PageService
from software_services.booking_services import BookingService
from software_services.examination_services import ExaminationService
from models.models import Status, db, User
from flask import request, abort
from models.models import Page
from parsers.facebook import parse_facebook_message
from platfroms.facebook_handler import FacebookHandler

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///medical_agent.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.getenv("SECRET_KEY")

login_manager = LoginManager(app)
db.init_app(app)
migrate = Migrate(app, db)

_seen_message_ids: set = set()
_seen_lock = threading.Lock()


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


@app.cli.command("create-admin")
@click.option("--username", prompt=True)
@click.option("--password", prompt=True)
def create_admin(username, password):
    existing = User.query.filter_by(username=username).first()
    if existing:
        print("Admin already exists")
        return
    admin = User(username=username, password=password)
    db.session.add(admin)
    db.session.commit()
    print("Admin created successfully")


#  AUTH

@app.route('/', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('users'))

    if request.method == 'POST':
        name     = request.form['name']
        password = request.form['password']
        user, message = UserService.login_user(name, password)

        if user:
            login_user(user)
            flash(message, 'success')
            return redirect(url_for('users'))
        else:
            flash(message, 'danger')

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('تم تسجيل الخروج بنجاح', 'success')
    return redirect(url_for('login'))


#  USERS

@app.route('/users')
@login_required
def users():
    all_users = UserService.get_all_users()
    return render_template('users.html', users=all_users)


@app.route('/users/new', methods=['GET', 'POST'])
@login_required
def create_user():
    if request.method == 'POST':
        name     = request.form['name']
        password = request.form['password']
        user, message = UserService.create_user(name, password)
        if user:
            flash(message, 'success')
            return redirect(url_for('users'))
        else:
            flash(message, 'danger')

    return render_template('create_user.html')


@app.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    user, message = UserService.get_user_by_id(user_id)

    if not user:
        flash(message, 'danger')
        return redirect(url_for('users'))

    if request.method == 'POST':
        name     = request.form['name']
        password = request.form['password']
        updated_user, message = UserService.update_user(user_id, name, password)

        if updated_user:
            flash(message, 'success')
            return redirect(url_for('users'))
        else:
            flash(message, 'danger')

    return render_template('edit_user.html', user=user)


#  CLINICS

@app.route('/clinics')
@login_required                              
def list_clinics():
    clinics, message = ClinicService.get_all_clinics()
    return render_template('clinics.html', clinics=clinics)


@app.route('/clinics/new', methods=['GET', 'POST'])
@login_required                              
def create_clinic():
    if request.method == 'POST':
        name    = request.form['name']
        address = request.form['address']
        info    = request.form['info']
        clinic, message = ClinicService.create_clinic(name, address, info)

        if clinic:
            flash(message, 'success')
            return redirect(url_for('list_clinics'))  
        else:
            flash(message, 'danger')

    return render_template('create_clinic.html')


@app.route('/clinics/<int:clinic_id>/edit', methods=['GET', 'POST'])   
@login_required                                                       
def edit_clinic(clinic_id):
    clinic, message = ClinicService.get_clinic_by_id(clinic_id)

    if not clinic:
        flash(message, 'danger')
        return redirect(url_for('list_clinics'))       

    if request.method == 'POST':
        name    = request.form['name']
        address = request.form['address']
        info    = request.form['info']
        updated_clinic, message = ClinicService.update_clinic(clinic_id, name, address, info)

        if updated_clinic:
            flash(message, 'success')
            return redirect(url_for('list_clinics'))   
        else:
            flash(message, 'danger')

    return render_template('edit_clinic.html', clinic=clinic)


#  DOCTORS

@app.route('/doctors')
@login_required
def list_doctors():
    doctors, message = DoctorService.get_all_doctors()
    return render_template('doctors.html', doctors=doctors)


@app.route('/doctors/new', methods=['GET', 'POST'])
@login_required
def create_doctor():
    clinics, _ = ClinicService.get_all_clinics()

    if request.method == 'POST':
        name = request.form['name']
        doctor_info = request.form['doctor_info']
        clinic_id = request.form['clinic_id']

        doctor, message = DoctorService.create_doctor(name, doctor_info, clinic_id)

        if doctor:
            flash(message, 'success')
            return redirect(url_for('list_doctors'))
        else:
            flash(message, 'danger')

    return render_template('create_doctor.html', clinics=clinics)


@app.route('/doctors/<int:doctor_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_doctor(doctor_id):
    doctor, message = DoctorService.get_doctor_by_id(doctor_id)

    if not doctor:
        flash(message, 'danger')
        return redirect(url_for('list_doctors'))       

    if request.method == 'POST':
        name        = request.form['name']
        doctor_info = request.form['doctor_info']
        updated_doctor, message = DoctorService.update_doctor(doctor_id, name, doctor_info)

        if updated_doctor:
            flash(message, 'success')
            return redirect(url_for('list_doctors'))   
        else:
            flash(message, 'danger')

    return render_template('edit_doctor.html', doctor=doctor)


@app.route('/doctors/<int:doctor_id>/delete', methods=['POST'])
@login_required
def delete_doctor(doctor_id):
    doctor, message = DoctorService.delete_doctor(doctor_id)

    if doctor:
        flash(message, 'success')
    else:
        flash(message, 'danger')

    return redirect(url_for('list_doctors'))


#  SPECIALTIES

@app.route('/specialties')
@login_required
def list_specialties():
    specialties, message = SpecialtyService.get_all_specialties()
    return render_template('specialties.html', specialties=specialties)


@app.route('/specialties/new', methods=['GET', 'POST'])
@login_required
def create_specialty():
    clinics, _ = ClinicService.get_all_clinics()

    if request.method == 'POST':
        name      = request.form['name']
        details   = request.form['details']
        clinic_id = request.form['clinic_id']

        specialty, message = SpecialtyService.create_specialty(name, details, clinic_id)

        if specialty:
            flash(message, 'success')
            return redirect(url_for('list_specialties'))
        else:
            flash(message, 'danger')

    return render_template('create_specialty.html', clinics=clinics)


@app.route('/specialties/<int:specialty_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_specialty(specialty_id):
    specialty, message = SpecialtyService.get_specialty_by_id(specialty_id)

    if not specialty:
        flash(message, 'danger')
        return redirect(url_for('list_specialties'))       

    if request.method == 'POST':
        name    = request.form['name']
        details = request.form['details']
        updated_specialty, message = SpecialtyService.update_specialty(specialty_id, name, details)

        if updated_specialty:
            flash(message, 'success')
            return redirect(url_for('list_specialties'))   
        else:
            flash(message, 'danger')

    return render_template('edit_specialty.html', specialty=specialty)


@app.route('/specialties/<int:specialty_id>/delete', methods=['POST'])
@login_required
def delete_specialty(specialty_id):
    specialty, message = SpecialtyService.delete_specialty(specialty_id)

    if specialty:
        flash(message, 'success')
    else:
        flash(message, 'danger')

    return redirect(url_for('list_specialties'))


#  SERVICES

@app.route('/services')
@login_required
def list_services():
    services, message = ServiceService.get_all_services()
    return render_template('services.html', services=services)


@app.route('/services/new', methods=['GET', 'POST'])
@login_required
def create_service():
    clinics, _ = ClinicService.get_all_clinics()

    if request.method == 'POST':
        name        = request.form['name']
        description = request.form['description']
        price       = request.form['price']
        clinic_id   = request.form['clinic_id']

        service, message = ServiceService.create_service(name, description, clinic_id, price)

        if service:
            flash(message, 'success')
            return redirect(url_for('list_services'))
        else:
            flash(message, 'danger')

    return render_template('create_service.html', clinics=clinics)


@app.route('/services/<int:service_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_service(service_id):
    service, message = ServiceService.get_service_by_id(service_id)

    if not service:
        flash(message, 'danger')
        return redirect(url_for('list_services'))

    if request.method == 'POST':
        name        = request.form['name']
        description = request.form['description']
        price       = request.form['price']

        updated_service, message = ServiceService.update_service(service_id, name, description, price)

        if updated_service:
            flash(message, 'success')
            return redirect(url_for('list_services'))
        else:
            flash(message, 'danger')

    return render_template('edit_service.html', service=service)


@app.route('/services/<int:service_id>/delete', methods=['POST'])
@login_required
def delete_service(service_id):
    service, message = ServiceService.delete_service(service_id)

    if service:
        flash(message, 'success')
    else:
        flash(message, 'danger')

    return redirect(url_for('list_services'))


#  PLATFORMS

@app.route('/platforms')
@login_required
def list_platforms():
    platforms, message = PlatformService.get_all_platforms()
    return render_template('platforms.html', platforms=platforms)


@app.route('/platforms/new', methods=['GET', 'POST'])
@login_required
def create_platform():
    if request.method == 'POST':
        name = request.form['name']
        platform, message = PlatformService.create_platform(name)
        if platform:
            flash(message, 'success')
            return redirect(url_for('list_platforms'))
        else:
            flash(message, 'danger')
    return render_template('create_platform.html')


@app.route('/platforms/<int:platform_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_platform(platform_id):
    platform, message = PlatformService.get_platform_by_id(platform_id)

    if not platform:
        flash(message, 'danger')
        return redirect(url_for('list_platforms'))

    if request.method == 'POST':
        name = request.form['name']
        updated_platform, message = PlatformService.update_platform(platform_id, name)

        if updated_platform:
            flash(message, 'success')
            return redirect(url_for('list_platforms'))
        else:
            flash(message, 'danger')

    return render_template('edit_platform.html', platform=platform)


#  PAGES

@app.route('/pages')
@login_required
def list_pages():
    pages, message = PageService.get_all_pages()
    return render_template('pages.html', pages=pages)


@app.route('/pages/new', methods=['GET', 'POST'])
@login_required
def create_page():
    if request.method == 'POST':
        clinic_id   = request.form['clinic_id']
        platform_id = request.form['platform_id']
        page_id     = request.form['page_id']
        token       = request.form['token']
        page, message = PageService.create_page(clinic_id, platform_id, page_id, token)
        if page:
            flash(message, 'success')
            return redirect(url_for('list_pages'))
        else:
            flash(message, 'danger')

    clinics, _   = ClinicService.get_all_clinics()
    platforms, _ = PlatformService.get_all_platforms()
    return render_template('create_page.html', clinics=clinics, platforms=platforms)


@app.route('/pages/<int:platform_id>/<page_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_page(platform_id, page_id):
    page, message = PageService.get_page_by_id(platform_id=platform_id, page_id=page_id)
    if not page:
        flash(message, 'danger')
        return redirect(url_for('list_pages'))

    if request.method == 'POST':
        clinic_id       = request.form['clinic_id']
        platform_id_new = request.form['platform_id']
        token           = request.form['token']
        updated_page, message = PageService.update_page(
            page_id=page.page_id,
            clinic_id=clinic_id,
            platform_id=platform_id_new,
            token=token,
        )
        if updated_page:
            flash(message, 'success')
            return redirect(url_for('list_pages'))
        else:
            flash(message, 'danger')

    clinics, _   = ClinicService.get_all_clinics()
    platforms, _ = PlatformService.get_all_platforms()
    return render_template('edit_page.html', page=page, clinics=clinics, platforms=platforms)


# BOOKINGS

@app.route('/bookings')
@login_required
def list_bookings():
    page      = request.args.get('page', 1, type=int)
    search    = request.args.get('search', '').strip()
    status    = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')

    pagination, message = BookingService.display_bookings(
        page=page, search=search, status=status,
        date_from=date_from, date_to=date_to
    )
    return render_template('bookings.html',
        pagination=pagination,
        bookings=pagination.items,
        search=search, status=status,
        date_from=date_from, date_to=date_to,
        Status=Status
    )


@app.route('/bookings/update_status/<int:booking_id>', methods=['POST'])
@login_required
def update_booking_status(booking_id):
    new_status_value = request.form.get('status_val')
    try:
        status_enum = Status(new_status_value)
    except ValueError:
        flash("حالة غير صحيحة", 'danger')
        return redirect(url_for('list_bookings'))

    updated, message = BookingService.update_booking_status(booking_id, status_enum)
    flash(message, 'success' if updated else 'danger')
    return redirect(url_for('list_bookings'))

@app.route('/bookings/export')
@login_required
def export_bookings():
    search    = request.args.get('search', '').strip()
    status    = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')

    # نفس الـ query بس بدون pagination - كل النتايج
    pagination, _ = BookingService.display_bookings(
        page=1, per_page=99999,
        search=search, status=status,
        date_from=date_from, date_to=date_to
    )
    bookings = pagination.items

    # عمل الـ Excel
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookings"

    # Headers
    ws.append(["الاسم", "التفاصيل", "التاريخ", "رقم الهاتف", "المنصة", "وقت الحجز", "الحالة"])

    # Data
    for b in bookings:
        ws.append([
            b.name,
            b.details,
            b.date,
            b.phone_number,
            b.comes_from,
            b.booking_time.strftime('%Y-%m-%d %H:%M'),
            b.status.value
        ])

    # حفظ وإرسال
    from io import BytesIO
    from flask import send_file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='bookings.xlsx'
    )

# COMPLAINTS

# COMPLAINTS
@app.route('/complaints')
@login_required
def list_complaints():
    page      = request.args.get('page', 1, type=int)
    search    = request.args.get('search', '').strip()
    status    = request.args.get('status', '')

    pagination, _ = ComplaintService.get_all_complaints(
        page=page, search=search, status=status,
    )
    return render_template('complaints.html',
        pagination=pagination,
        complaints=pagination.items,
        search=search, status=status,
    )


@app.route('/complaints/update_status/<int:complaint_id>', methods=['POST'])
@login_required
def update_complaint_status(complaint_id):
    new_status_value = request.form.get('status_val')
    try:
        status_enum = Status(new_status_value)
    except ValueError:
        flash("حالة غير صحيحة", 'danger')
        return redirect(url_for('list_complaints'))
    updated, message = ComplaintService.update_complaint_status(complaint_id, status_enum)
    flash(message, 'success' if updated else 'danger')
    return redirect(url_for('list_complaints'))

@app.route('/complaints/export')
@login_required
def export_complaints():
    search = request.args.get('search', '').strip()
    status = request.args.get('status', '')

    pagination, _ = ComplaintService.get_all_complaints(
        page=1, per_page=99999,
        search=search, status=status
    )
    complaints = pagination.items

    import openpyxl
    from io import BytesIO
    from flask import send_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Complaints"
    ws.append(["رقم الهاتف", "الشكوى", "تاريخ الإنشاء", "المنصة", "الحالة"])

    for c in complaints:
        ws.append([
            c.phone_number,
            c.complaint_text,
            c.created_at.strftime('%Y-%m-%d %H:%M'),
            c.comes_from or 'غير محدد',
            c.status.value if c.status else 'Pending'
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='complaints.xlsx'
    )

# EXAMINATIONS
@app.route('/examinations')
@login_required
def list_examinations():
    page      = request.args.get('page', 1, type=int)
    search    = request.args.get('search', '').strip()
    status    = request.args.get('status', '')
    

    pagination, _ = ExaminationService.get_all_examinations(
        page=page, search=search, status=status,
    
    )
    return render_template('examinations.html',
        pagination=pagination,
        examinations=pagination.items,
        search=search, status=status,
    )


@app.route('/examinations/update_status/<int:examination_id>', methods=['POST'])
@login_required
def update_examination_status(examination_id):
    new_status_value = request.form.get('status_val')
    try:
        status_enum = Status(new_status_value)
    except ValueError:
        flash("حالة غير صحيحة", 'danger')
        return redirect(url_for('list_examinations'))
    updated, message = ExaminationService.update_examination_status(examination_id, status_enum)
    flash(message, 'success' if updated else 'danger')
    return redirect(url_for('list_examinations'))


@app.route('/examinations/export')
@login_required
def export_examinations():
    search = request.args.get('search', '').strip()
    status = request.args.get('status', '')

    pagination, _ = ExaminationService.get_all_examinations(
        page=1, per_page=99999,
        search=search, status=status
    )
    examinations = pagination.items

    import openpyxl
    from io import BytesIO
    from flask import send_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Examinations"
    ws.append(["رقم الهاتف", "الأعراض", "تاريخ الإنشاء", "المنصة", "الحالة"])

    for e in examinations:
        ws.append([
            e.phone_number,
            e.symptom_text,
            e.created_at.strftime('%Y-%m-%d %H:%M'),
            e.comes_from or 'غير محدد',
            e.status.value if e.status else 'Pending'
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='examinations.xlsx'
    )
# ─────────────────────────────────────
# WEBHOOK
# ─────────────────────────────────────

VERIFY_TOKEN = os.getenv("VERIFY_TOKEN")


@app.route("/webhook/facebook", methods=["GET", "POST"])
def fb_webhook():

    if request.method == "GET":
        if request.args.get("hub.verify_token") == VERIFY_TOKEN:
            return request.args.get("hub.challenge", "")
        abort(403)

    try:
        payload = request.json or {}
        entries = payload.get("entry", [])
    except Exception:
        return "OK", 200

    def process():
        for entry in entries:
            page_id = entry.get("id")
            if not page_id:
                continue

            with app.app_context():
                try:
                    page = Page.query.filter_by(page_id=page_id).first()
                    if not page:
                        continue

                    handler = FacebookHandler(page)

                    for messaging in entry.get("messaging", []):

                        mid = messaging.get("message", {}).get("mid")
                        if mid:
                            with _seen_lock:
                                if mid in _seen_message_ids:
                                    continue
                                _seen_message_ids.add(mid)
                                if len(_seen_message_ids) > 10_000:
                                    _seen_message_ids.clear()

                        message = parse_facebook_message(
                            messaging=messaging,
                            page_id=page.page_id,
                            platform_id=handler.platform_id,
                            platform_name=handler.platform_name,
                        )

                        if not message:
                            continue

                        reply = handler.handle(message)

                        if reply:
                            handler.send(message.sender_id, reply)

                except Exception:
                    import traceback
                    print("WEBHOOK THREAD ERROR:")
                    print(traceback.format_exc())

    threading.Thread(target=process, daemon=True).start()
    return "OK", 200


if __name__ == '__main__':
    app.run()